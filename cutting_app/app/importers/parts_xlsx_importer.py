from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from unicodedata import normalize
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from cutting_app.app.importers.parts_table_importer import EditablePartRow


@dataclass(frozen=True)
class PartsXlsxImportError:
	message: str
	excel_row_number: int | None = None
	table_row_number: int | None = None


@dataclass(frozen=True)
class PartsXlsxImportResult:
	rows: list[EditablePartRow]
	errors: list[PartsXlsxImportError]
	filename: str
	sheet_name: str
	skipped_row_count: int


_HEADER_SCAN_LIMIT = 20
_REQUIRED_COLUMNS = ("number", "l_mm", "w_mm", "quantity")
_COLUMN_LABELS = {
	"number": "Позиция / Номер",
	"l_mm": "Длинна / Длина / L",
	"w_mm": "Ширина / W",
	"quantity": "Колличество / Количество",
}
_COLUMN_ALIASES = {
	"cut": ("кроить",),
	"number": ("позиция", "номер", "номердетали", "№"),
	"name": ("наименования", "наименование", "название"),
	"l_mm": ("длинна", "длина", "l", "length"),
	"w_mm": ("ширина", "w", "width"),
	"quantity": ("колличество", "количество", "колво", "quantity"),
	"rotation": ("ориентация", "поворот", "разрешитьповорот"),
	"L1": ("l1", "l1обозн", "l1обозначение"),
	"L2": ("l2", "l2обозн", "l2обозначение"),
	"W1": ("w1", "w1обозн", "w1обозначение"),
	"W2": ("w2", "w2обозн", "w2обозначение"),
}


def import_parts_xlsx(
	contents: bytes,
	*,
	filename: str = "",
) -> PartsXlsxImportResult:
	display_filename = Path(filename).name if filename else ""
	try:
		workbook = load_workbook(
			BytesIO(contents),
			data_only=True,
			keep_links=False,
			read_only=True,
		)
	except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError):
		return _error_result(
			filename=display_filename,
			message="Файл XLSX не удалось открыть. Проверь, что выбран настоящий файл Excel формата .xlsx.",
		)

	try:
		worksheet = workbook.active
		header_row_number, columns = _find_header_row(worksheet)
		if header_row_number is None:
			return _missing_headers_result(
				filename=display_filename,
				sheet_name=worksheet.title,
				columns=columns,
			)

		rows: list[EditablePartRow] = []
		errors: list[PartsXlsxImportError] = []
		skipped_row_count = 0
		for excel_row_number, values in enumerate(
			worksheet.iter_rows(
				min_row=header_row_number + 1,
				values_only=True,
			),
			start=header_row_number + 1,
		):
			if _is_empty_data_row(values, columns):
				continue

			should_import, cut_error = _parse_cut_value(
				_value_at(values, columns.get("cut"))
			)
			if not should_import and cut_error is None:
				skipped_row_count += 1
				continue

			table_row_number = len(rows) + 1
			row, row_errors = _build_editable_row(values, columns)
			rows.append(row)
			if cut_error:
				row_errors.insert(0, cut_error)
			if row_errors:
				errors.append(
					PartsXlsxImportError(
						message=" ".join(row_errors),
						excel_row_number=excel_row_number,
						table_row_number=table_row_number,
					)
				)

		if not rows and not errors:
			errors.append(
				PartsXlsxImportError(
					message="В Excel не найдено ни одной строки деталей для раскроя."
				)
			)

		return PartsXlsxImportResult(
			rows=rows,
			errors=errors,
			filename=display_filename,
			sheet_name=worksheet.title,
			skipped_row_count=skipped_row_count,
		)
	finally:
		workbook.close()


def _find_header_row(worksheet: object) -> tuple[int | None, dict[str, int]]:
	best_columns: dict[str, int] = {}
	for row_number, values in enumerate(
		worksheet.iter_rows(
			min_row=1,
			max_row=min(worksheet.max_row, _HEADER_SCAN_LIMIT),
			values_only=True,
		),
		start=1,
	):
		columns = _map_columns(values)
		if len(columns) > len(best_columns):
			best_columns = columns
		if all(column in columns for column in _REQUIRED_COLUMNS):
			return row_number, columns

	return None, best_columns


def _map_columns(values: tuple[object, ...]) -> dict[str, int]:
	columns: dict[str, int] = {}
	for index, value in enumerate(values):
		header = _normalize_header(value)
		if not header:
			continue
		for column_name, aliases in _COLUMN_ALIASES.items():
			if column_name in columns:
				continue
			if header in {_normalize_header(alias) for alias in aliases}:
				columns[column_name] = index
				break
	return columns


def _missing_headers_result(
	*,
	filename: str,
	sheet_name: str,
	columns: dict[str, int],
) -> PartsXlsxImportResult:
	missing = [
		_COLUMN_LABELS[column]
		for column in _REQUIRED_COLUMNS
		if column not in columns
	]
	return PartsXlsxImportResult(
		rows=[],
		errors=[
			PartsXlsxImportError(
				message="В Excel не найдены обязательные колонки: " + ", ".join(missing) + "."
			)
		],
		filename=filename,
		sheet_name=sheet_name,
		skipped_row_count=0,
	)


def _build_editable_row(
	values: tuple[object, ...],
	columns: dict[str, int],
) -> tuple[EditablePartRow, list[str]]:
	number = _display_value(_value_at(values, columns["number"]))
	name = _display_value(_value_at(values, columns.get("name")))
	if not name and number:
		name = f"Позиция {number}"

	rotation_allowed, rotation_error = _parse_rotation_value(
		_value_at(values, columns.get("rotation"))
	)
	edge_values: dict[str, bool] = {}
	errors = [rotation_error] if rotation_error else []
	for side in ("L1", "L2", "W1", "W2"):
		has_edge, edge_error = _parse_edge_marker(
			_value_at(values, columns.get(side)),
			side,
		)
		edge_values[side] = has_edge
		if edge_error:
			errors.append(edge_error)

	return (
		EditablePartRow(
			number=number,
			name=name,
			l_mm=_display_value(_value_at(values, columns["l_mm"])),
			w_mm=_display_value(_value_at(values, columns["w_mm"])),
			quantity=_display_value(_value_at(values, columns["quantity"])),
			rotation_allowed=rotation_allowed,
			L1=edge_values["L1"],
			L2=edge_values["L2"],
			W1=edge_values["W1"],
			W2=edge_values["W2"],
		),
		errors,
	)


def _parse_cut_value(value: object) -> tuple[bool, str | None]:
	normalized = _normalize_value(value)
	if normalized in ("", "да", "yes", "1", "+", "*"):
		return True, None
	if normalized in ("нет", "no", "0", "-"):
		return False, None
	return (
		True,
		f"Значение «Кроить» ({_display_value(value)}) не распознано; строка оставлена для ручной проверки.",
	)


def _parse_rotation_value(value: object) -> tuple[bool, str | None]:
	normalized = _normalize_value(value)
	if normalized in (
		"",
		"не задана",
		"не задано",
		"да",
		"разрешена",
		"разрешен",
		"с поворотом",
	):
		return True, None
	if normalized in (
		"нет",
		"без поворота",
		"запрещена",
		"запрещен",
	):
		return False, None
	return (
		True,
		f"Ориентация «{_display_value(value)}» не распознана; проверь разрешение поворота.",
	)


def _parse_edge_marker(value: object, side: str) -> tuple[bool, str | None]:
	normalized = _normalize_value(value)
	if normalized in ("", "-", "—", "нет", "0"):
		return False, None
	if normalized in ("*", "+", "да", "1", "x", "х"):
		return True, None
	return (
		False,
		f"Кромка {side}: отметка «{_display_value(value)}» не распознана; используй «*» или пустую ячейку.",
	)


def _is_empty_data_row(
	values: tuple[object, ...],
	columns: dict[str, int],
) -> bool:
	return all(
		_normalize_value(_value_at(values, index)) == ""
		for index in columns.values()
	)


def _value_at(values: tuple[object, ...], index: int | None) -> object:
	if index is None or index >= len(values):
		return None
	return values[index]


def _display_value(value: object) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "1" if value else "0"
	if isinstance(value, float) and value.is_integer():
		return str(int(value))
	return str(value).strip()


def _normalize_value(value: object) -> str:
	return _display_value(value).strip().lower().replace("ё", "е")


def _normalize_header(value: object) -> str:
	text = normalize("NFKC", _display_value(value)).lower().replace("ё", "е")
	text = text.replace("№", "номер")
	return "".join(character for character in text if character.isalnum())


def _error_result(*, filename: str, message: str) -> PartsXlsxImportResult:
	return PartsXlsxImportResult(
		rows=[],
		errors=[PartsXlsxImportError(message=message)],
		filename=filename,
		sheet_name="",
		skipped_row_count=0,
	)
